"""Reusable Excel export helpers for future notebook refactoring."""

from __future__ import annotations

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def write_workbook_tables(summary, diagnostics, workbook_path):
    """Write summary and diagnostics tables to an Excel workbook."""
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary table", index=True)
        diagnostics.to_excel(writer, sheet_name="Sheet diagnostics", index=False)


def rebuild_summary_header(worksheet):
    """Rebuild merged top-level headers after pandas Excel export."""
    worksheet.delete_rows(3)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    worksheet.cell(row=1, column=1).value = "Variable"
    worksheet.cell(row=1, column=2).value = "Category"

    current_column = 3
    while current_column <= worksheet.max_column:
        group_name = worksheet.cell(row=1, column=current_column).value
        if group_name:
            start_column = current_column
            end_column = current_column
            while (
                end_column + 1 <= worksheet.max_column
                and worksheet.cell(row=1, column=end_column + 1).value in [None, ""]
            ):
                end_column += 1
            worksheet.merge_cells(
                start_row=1,
                start_column=start_column,
                end_row=1,
                end_column=end_column,
            )
            current_column = end_column + 1
        else:
            current_column += 1


def apply_summary_sheet_style(
    worksheet,
    section_labels,
    not_applicable_marker="--",
    column_widths=None,
):
    """Apply workbook styling used by the summary table sheet."""
    header_fill_top = PatternFill("solid", fgColor="B7DEE8")
    header_fill_bottom = PatternFill("solid", fgColor="E2F0D9")
    section_fill = PatternFill("solid", fgColor="E2F0D9")
    not_applicable_fill = PatternFill("solid", fgColor="D9D9D9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    cell_border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row in [1, 2]:
                cell.font = Font(bold=True)
                cell.fill = header_fill_top if cell.row == 1 else header_fill_bottom

    for row_number in range(3, worksheet.max_row + 1):
        variable_cell = worksheet.cell(row=row_number, column=1)
        category_cell = worksheet.cell(row=row_number, column=2)
        variable_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        category_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        variable_cell.font = Font(bold=True)

        is_section_row = variable_cell.value in section_labels and not category_cell.value
        for column_number in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_number, column=column_number)
            if cell.value == not_applicable_marker:
                cell.fill = not_applicable_fill
                cell.font = Font(color="666666")
            elif is_section_row:
                cell.fill = section_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = white_fill

    for column_letter, width in (column_widths or {}).items():
        worksheet.column_dimensions[column_letter].width = width

    worksheet.freeze_panes = "C3"
    worksheet.sheet_view.showGridLines = False


def export_styled_summary_workbook(summary, diagnostics, workbook_path, section_labels, column_widths):
    """Export the styled workbook used for reporting."""
    write_workbook_tables(summary, diagnostics, workbook_path)
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook["Summary table"]
    rebuild_summary_header(worksheet)
    apply_summary_sheet_style(worksheet, section_labels=section_labels, column_widths=column_widths)
    workbook.save(workbook_path)
